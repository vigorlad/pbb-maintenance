"""
인천국제공항 항공기 운항 현황 상세 조회 → 엑셀 추출
T1(제1터미널), 탑승동, T2(제2터미널) 게이트 이용 항공편 정보
"""
import os
import sys
import requests
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# ── 설정 ──────────────────────────────────────────────
SERVICE_KEY = os.environ.get("SERVICE_KEY", "")
BASE_URL = "https://apis.data.go.kr/B551177/statusOfAllFltDeOdp"
NUM_OF_ROWS = 1000

TERMINAL_MAP = {
    "P01": "T1(제1터미널)",
    "P02": "탑승동",
    "P03": "T2(제2터미널)",
}
PASSENGER_TERMINALS = set(TERMINAL_MAP.keys())


# ── API 호출 ──────────────────────────────────────────
def fetch_all_flights(operation, search_date):
    """출발/도착 오퍼레이션의 전체 데이터를 페이징 순회하여 수집"""
    url = f"{BASE_URL}/{operation}"
    all_items = []
    page = 1

    while True:
        params = {
            "serviceKey": SERVICE_KEY,
            "type": "json",
            "numOfRows": NUM_OF_ROWS,
            "pageNo": page,
            "searchDate": search_date,
        }
        print(f"  [{operation}] 페이지 {page} 요청 중...")
        resp = requests.get(url, params=params, timeout=30)
        resp.raise_for_status()
        data = resp.json()

        body = data.get("response", {}).get("body", {})
        total_count = body.get("totalCount", 0)
        items = body.get("items", [])

        if not items:
            break

        all_items.extend(items)
        print(f"  [{operation}] 페이지 {page}: {len(items)}건 수신 (누적 {len(all_items)}/{total_count})")

        if len(all_items) >= total_count:
            break
        page += 1

    return all_items



# ── 날짜/시간 포맷 ───────────────────────────────────────
def fmt_date(raw):
    """'202412220005' → '2024-12-22'"""
    if not raw or raw == "-":
        return "-"
    try:
        dt = datetime.strptime(raw.strip(), "%Y%m%d%H%M")
        return dt.strftime("%Y-%m-%d")
    except ValueError:
        return raw


def fmt_time(raw):
    """'202412220005' → '00:05'"""
    if not raw or raw == "-":
        return "-"
    try:
        dt = datetime.strptime(raw.strip(), "%Y%m%d%H%M")
        return dt.strftime("%H%M")
    except ValueError:
        return raw


# ── 엑셀 작성 ─────────────────────────────────────────
COLUMNS = [
    ("운항일자", "_date"),
    ("출도착", "_flight_type"),
    ("편명", "flightId"),
    ("STA/STD", "scheduleDatetime"),
    ("등록기호", "aircraftRegNo"),
    ("ATA/ATD", "estimatedDatetime"),
    ("운항여부", "remark"),
    ("주기장", "fstandPosition"),
    ("기종", "aircraftSubtype"),
    ("출발지공항명", "_dep_airport"),
    ("도착지공항명", "_arr_airport"),
]

HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(name="맑은 고딕", bold=True, color="FFFFFF", size=10)
CELL_FONT = Font(name="맑은 고딕", size=10)
THIN_BORDER = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)


def write_sheet(wb, sheet_name, all_items):
    """워크북에 시트를 추가하고 데이터를 기록. all_items는 (item, flight_type) 튜플 리스트"""
    ws = wb.create_sheet(title=sheet_name)

    # 헤더
    for col_idx, (col_name, _) in enumerate(COLUMNS, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER

    # 예정일시 기준 정렬
    all_items.sort(key=lambda x: x[0].get("scheduleDatetime", "") or "")

    # 데이터
    for row_idx, (item, flight_type) in enumerate(all_items, 2):
        for col_idx, (_, field) in enumerate(COLUMNS, 1):
            if field == "_date":
                value = fmt_date(str(item.get("scheduleDatetime", "") or ""))
            elif field == "_flight_type":
                value = flight_type
            elif field in ("scheduleDatetime", "estimatedDatetime"):
                value = fmt_time(str(item.get(field, "") or ""))
            elif field == "_dep_airport":
                value = item.get("airport", "-") if flight_type == "A" else "-"
            elif field == "_arr_airport":
                value = item.get("airport", "-") if flight_type == "D" else "-"
            else:
                value = item.get(field, "-") or "-"
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = CELL_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = THIN_BORDER

    # 컬럼 너비 자동 조절
    for col_idx in range(1, len(COLUMNS) + 1):
        max_len = 0
        for row in ws.iter_rows(min_col=col_idx, max_col=col_idx, values_only=True):
            val = str(row[0]) if row[0] else ""
            length = sum(2 if ord(c) > 127 else 1 for c in val)
            max_len = max(max_len, length)
        ws.column_dimensions[get_column_letter(col_idx)].width = max(max_len + 3, 10)

    # 필터 설정
    ws.auto_filter.ref = ws.dimensions


def date_range(start, end):
    """start~end 사이 날짜 문자열(YYYYMMDD) 리스트 반환"""
    s = datetime.strptime(start, "%Y%m%d")
    e = datetime.strptime(end, "%Y%m%d")
    dates = []
    while s <= e:
        dates.append(s.strftime("%Y%m%d"))
        s += timedelta(days=1)
    return dates


def main():
    today = datetime.now()
    min_date = (today + timedelta(days=-3)).strftime("%Y%m%d")
    max_date = (today + timedelta(days=6)).strftime("%Y%m%d")

    # 인자 파싱: 시작일 [종료일]
    if len(sys.argv) >= 3:
        start_date = sys.argv[1]
        end_date = sys.argv[2]
    elif len(sys.argv) == 2:
        start_date = sys.argv[1]
        end_date = sys.argv[1]
    else:
        start_date = today.strftime("%Y%m%d")
        end_date = start_date

    # 범위 검증
    if start_date < min_date or end_date > max_date:
        print(f"조회 가능 범위: {min_date} ~ {max_date} (오늘 기준 -3일 ~ +6일)")
        print(f"입력한 범위: {start_date} ~ {end_date}")
        sys.exit(1)

    dates = date_range(start_date, end_date)
    print(f"=== 인천국제공항 항공기 운항 현황 조회 ({start_date}~{end_date}, {len(dates)}일) ===\n")

    # 1) 날짜별로 출발/도착 데이터 수집
    departures = []
    arrivals = []
    for d in dates:
        print(f"[{d} 출발편 조회]")
        departures.extend(fetch_all_flights("getFltDeparturesDeOdp", d))
        print(f"[{d} 도착편 조회]")
        arrivals.extend(fetch_all_flights("getFltArrivalsDeOdp", d))
        print()

    # 2) 터미널별로 출도착 합쳐서 분류
    terminal_items = {tid: [] for tid in PASSENGER_TERMINALS}
    for item in departures:
        tid = item.get("terminalId", "")
        if tid in PASSENGER_TERMINALS:
            terminal_items[tid].append((item, "D"))
    for item in arrivals:
        tid = item.get("terminalId", "")
        if tid in PASSENGER_TERMINALS:
            terminal_items[tid].append((item, "A"))

    # 3) 엑셀 생성 (터미널별 시트)
    wb = Workbook()
    wb.remove(wb.active)

    sheet_order = [
        ("P01", "T1"),
        ("P02", "탑승동"),
        ("P03", "T2"),
    ]
    for tid, sheet_name in sheet_order:
        items = terminal_items[tid]
        write_sheet(wb, sheet_name, items)
        print(f"  시트 [{sheet_name}]: {len(items)}건")

    # 4) 파일 저장
    if start_date == end_date:
        filename = f"인천공항_운항현황_{start_date}.xlsx"
    else:
        filename = f"인천공항_운항현황_{start_date}_{end_date}.xlsx"
    wb.save(filename)
    print(f"\n엑셀 파일 저장 완료: {filename}")
    print(f"총 출발편: {len(departures)}건, 도착편: {len(arrivals)}건")


if __name__ == "__main__":
    main()
